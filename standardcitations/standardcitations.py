# -*- coding: utf-8 -*-
# @Author: Martin Isaksson
# @Date:   2019-01-14 08:38:09
# @Last Modified by:   Martin Isaksson
# @Last Modified time: 2019-02-06 20:13:06

import argparse
from argparse import RawTextHelpFormatter
from openpyxl import load_workbook
from bibtexparser.bwriter import BibTexWriter
from bibtexparser.bibdatabase import BibDatabase
import sys
from lxml import html
import requests
from tqdm import tqdm


DESCRIPTION = """
3GPP Bibtex entry generator --- Convert 3GPP document list from .xls to .bib.

1. Go to the [3GPP Portal](https://portal.3gpp.org/#55936-specifications)
2. Generate the list of specifications you want.
3. Download to Excel and save file
4. Run
   `python 3gpp-citations.py -i exported.xlsx -o 3gpp.bib`
5. Use in LaTeX.

* The output `bibtex` class is set to `@techreport`.
* The version and date are read from the URL, but it is slow so it takes
  a while to parse the list. If you find an easy solution to this, let me know.
"""

EPILOG = """

Example output:

@Techreport{3gpp.36.331,
  author = "3GPP",
  title = "{Evolved Universal Terrestrial Radio Access (E-UTRA);
        Radio Resource Control (RRC); Protocol specification}",
  type = "TS",
  institution = "{3rd Generation Partnership Project (3GPP)}",
  number = "{36.331}",
  days = 11,
  month = jul,
  year = 2016,
  url = "http://www.3gpp.org/dynareport/36331.htm",
}
"""


def main(args):
    """
    The main function that does all the heavy lifting.
    """

    db = BibDatabase()
    db.entries = []

    wb2 = load_workbook(args.input)
    ws = wb2[wb2.sheetnames[0]]

    row_count = ws.max_row - 1  # Skipping header

    # Iterate over the rows in the Excel-sheet but skip the header.

    for row in tqdm(
            ws.iter_rows(row_offset=1),
            total=row_count):

        number = row[0].value
        title = row[2].value
        type = row[1].value

        if number is None:
            continue

        if args.xelatex:
            url = "http://www.3gpp.org/\-DynaReport/\-{}.htm".format(
                number.replace(".", ""))
        else:
            url = "http://www.3gpp.org/DynaReport/{}.htm".format(
                number.replace(".", ""))

        entry = {
            'ID': "3gpp.{}".format(number),
            'ENTRYTYPE': "techreport",
            'title': "{{{}}}".format(title),
            'type': type,
            'author': "3GPP",
            'institution': "{3rd Generation Partnership Project (3GPP)}",
            'number': number,
            'url': url
        }

        if row[0].hyperlink is not None:
            # entry['url'] = row[0].hyperlink.target

            page = requests.get(row[0].hyperlink.target)
            tree = html.fromstring(page.content)

            for release in range(2):

                row = tree.xpath(
                    ('//tr[@id="SpecificationReleaseControl1_rpbReleases_i{}'
                        '_ctl00_specificationsVersionGrid_ctl00__0"]/td/div/a')
                    .format(release))

                if len(row) > 0:
                    daterow = tree.xpath(
                        ('//tr[@id="SpecificationReleaseControl1_rpbReleases'
                         '_i{}_ctl00_specificationsVersionGrid_ctl00__0"]/td')
                        .format(release))

                    entry['note'] = "Version {}".format(row[1].text.strip())

                    if daterow[2].text.strip() is not "":
                        date = daterow[2].text.split('-')
                        entry['day'] = date[2].strip()
                        entry['year'] = date[0].strip()
                        entry['month'] = date[1].strip()
                    break

        db.entries.append(entry)

    writer = BibTexWriter()

    if args.output is not None:
        with open(args.output, 'w') as bibfile:
            bibfile.write(writer.write(db))
    else:
        print(writer.write(db))


def parse_args(args):
    """
    Parse arguments
    """

    parser = argparse.ArgumentParser(
        description=DESCRIPTION,
        epilog=EPILOG,
        formatter_class=RawTextHelpFormatter)

    parser.add_argument('--input', '-i', metavar='INPUT',
                        required=True,
                        help=('The Excel file generated by and '
                              'exported from the 3GPP Portal '
                              '(https://portal.3gpp.org)'))
    parser.add_argument('--output', '-o', metavar='OUTPUT',
                        help=('The bib file to write to. '
                              'STDOUT is used if omitted.'))
    parser.add_argument('--xelatex',
                        action='store_true',
                        help='Use line breaks')

    args = parser.parse_args(args)

    return args


if __name__ == "__main__":
    main(parse_args(sys.argv[1:]))
